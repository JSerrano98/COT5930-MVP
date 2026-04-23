"""
Session Manager:
- Listens for LabStreaming Layer (LSL) outlets and creates inlets
- Broadcast stream data to connected WebSocket clients in real-time
- Provides an HTTP endpoint to list available streams and their metadata
- Manages client connections and disconnections

Program runs when the user starts a new session from the frontend. 
Just a simple FastAPI app with a background task to read from LSL inlets 
and broadcast to WebSocket clients.
"""

import asyncio
import json
import openpyxl
import pandas as pd
from pathlib import Path
from contextlib import asynccontextmanager

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from pylsl import StreamInlet, resolve_streams

inlets: list[dict] = []  # {"name": str, "type": str, "inlet": StreamInlet}
clients: list[WebSocket] = []
cols = ['TimeStamp', 'StreamName', 'data']
wb = openpyxl.Workbook()
wb.active.append(cols)

def create_path():
    script_dir = Path(__file__).parent
    path = script_dir / 'CSV'
    return str(path)


def Excel_payload(timestamp,streamname, data ):
    wb.active.append([timestamp, streamname, data])
    last_row = wb.active.max_row
    wb.active.cell(row=last_row, column=3).number_format = '0.000000'
    wb.save(create_path() + '/data2.xlsx')
    return

def discover_streams(timeout: float = 3.0) -> list[dict]:
    """
    Find all LSL outlets on the network and create an inlet for each.
    Args:
        timeout (float): How long to wait for streams to appear in seconds.
    Returns:
        List of dicts containing stream metadata and inlet object.
    """
    print(f"Scanning for LSL streams ({timeout}s)...")
    found = resolve_streams(timeout)
    print(f"Found {len(found)} stream(s)")

    result = []
    for info in found:
        inlet = StreamInlet(info, max_buflen=1)
        inlet.open_stream()
        result.append({
            "name": info.name(),
            "type": info.type(),
            "channels": info.channel_count(),
            "rate": info.nominal_srate(),
            "inlet": inlet,
        })
        print(f"\n SIG: {info.name()} | {info.type()} | {info.channel_count()}ch @ {info.nominal_srate()}Hz")

    return result


async def read_lsl_loop(update_rate: float = 250.0):
    """
    Pull samples from all inlets and broadcast to all WebSocket clients in "real-time".
    
    Args:
        update_rate (float): The rate at which to update clients in Hz. 
            This is a throttle to prevent CPU overload, 
            not the actual sample rate of the streams.
            Default is 250Hz, or 4ms between updates, 
            which should be sufficient for most biosignals.
    """
    
    while True:
        for stream in inlets:
            inlet: StreamInlet = stream["inlet"]
            
            # pull_sample with timeout=0.0 is non-blocking
            sample, timestamp = inlet.pull_sample(timeout=0.0)
            if sample is not None:
                Excel_payload(timestamp, stream["name"], sample[0])
                payload = json.dumps({
                    "stream": stream["name"],
                    "type": stream["type"],
                    "timestamp": timestamp,
                    "data": sample,
                })

                # broadcast to every connected client
                disconnected = []
                for ws in clients:
                    try:
                        await ws.send_text(payload)
                    except Exception:
                        disconnected.append(ws)
                for ws in disconnected:
                    clients.remove(ws)

        # CPU throttle 
        await asyncio.sleep(1.0 / update_rate) 
        # TODO: !OutOfScope, add frontend absraction to allow user to configure update rate at session start 


@asynccontextmanager
async def lifespan(app: FastAPI):
    global inlets
    inlets = discover_streams(timeout=3.0)
    task = asyncio.create_task(read_lsl_loop())
    yield
    task.cancel()
    for stream in inlets:
        stream["inlet"].close_stream()


app = FastAPI(lifespan=lifespan)


# ════════════════════════════════════════════════════════════════════
# ROUTES
# ════════════════════════════════════════════════════════════════════
@app.get("/streams")
def list_streams():
    """
    This is used by the frontend to display available streams and their properties.
    Return metadata about discovered LSL streams.

    Returns:
        List of dicts with stream name, type, channel count, and sample rate.
    """
    return [
        {
            "name": s["name"],
            "type": s["type"],
            "channels": s["channels"],
            "rate": s["rate"],
        }
        for s in inlets
    ]

# ════════════════════════════════════════════════════════════════════
# WEBSOCKET ENDPOINT
# ════════════════════════════════════════════════════════════════════
@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws.accept()
    clients.append(ws)
    print(f"Client connected ({len(clients)} total)")
    try:
        # keep the connection alive, wait for client disconnect
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        clients.remove(ws)
        print(f"Client disconnected ({len(clients)} total)")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
